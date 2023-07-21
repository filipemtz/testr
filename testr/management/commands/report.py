
import uuid
import random
import re
from time import sleep
from django.core.management.base import BaseCommand
from testr.autojudge.autojudge_runner import AutoJudgeRunner
from testr.models.submission import Submission, SubmissionStatus
from multiprocessing import Pool
from openpyxl import Workbook

import os
import shutil
from tqdm import tqdm
from testr.models import *


class ReportExporter:
    def __init__(self, output_path) -> None:
        self._output_path = output_path

    def export(self):
        self._report_dir = f"reports/{uuid.uuid4()}"

        # create dir and xlsx file for each courses
        self._create_course_reports()

        # zip the dir and save it in the output_path
        self._compress_and_remove_report_dir()

    @staticmethod
    def _sanitize(filename):
        return re.sub(r'\W+', '_', filename)

    def _create_course_reports(self):
        courses = Course.objects.all()

        for course in tqdm(courses):
            # create dir for the course and xlsx file to store the report
            course_name = ReportExporter._sanitize(
                f"{course.name}_{course.id}")
            course_dir = f"{self._report_dir}/{course_name}"
            os.makedirs(course_dir, exist_ok=True)

            report = Workbook()
            sheet = report.active
            sheet.title = "Summary"

            # write the names of the enrolled students in the first column of the xlsx file
            enrolls = Enrollment.objects.filter(
                course=course).order_by('enrolled_at')

            sheet.cell(row=1, column=1, value='Id')
            sheet.cell(row=1, column=2, value='Name')
            student_rows = {}
            for idx, enroll in enumerate(enrolls):
                name = f"{enroll.student.first_name} {enroll.student.last_name}"
                student_row = idx + 2
                sheet.cell(row=student_row, column=1, value=enroll.student.id)
                sheet.cell(row=student_row, column=2, value=name)
                student_rows[enroll.student.id] = student_row

            # create dir for each question and name the colums with the question names
            for question_idx, question in enumerate(Question.objects.filter(section__course=course)):
                question_name = f"{question.name}_{question.id}"

                question_dir = ReportExporter._sanitize(question_name)
                question_dir = f"{course_dir}/{question_dir}"
                os.mkdir(question_dir)

                question_col = question_idx + 3
                sheet.cell(row=1, column=question_col, value=question_name)

                # save most recent submission for each student in each question
                # submissions are sorted so that the most recent comes first
                submissions = list(
                    Submission.objects.filter(question=question).order_by('-created_at'))

                # save the most recent submission for each student and
                # update the xlsx file with the results of the submissions
                saved_students = set()
                for submission in submissions:
                    student = submission.student

                    if student.id not in saved_students:
                        # if student unrolled after the submission or user is a teacher
                        if student.id not in student_rows:
                            name = f"{student.first_name} {student.last_name}"
                            student_row = len(student_rows) + 2
                            sheet.cell(row=student_row, column=1,
                                       value=student.id)
                            sheet.cell(row=student_row, column=2, value=name)
                            student_rows[student.id] = student_row
                            # print(
                            #    f"Warning: student '{name}' is not enrolled in course '{course.name}', but has submitted a solution for question '{question.name}'.")

                        row = student_rows[student.id]
                        val = 1 if submission.status == SubmissionStatus.SUCCESS else 0
                        sheet.cell(row=row, column=question_col, value=val)

                        student_name = ReportExporter._sanitize(
                            f"{student.first_name} {student.last_name}")
                        submission_name = f"{student_name}_{student.id}_{submission.status}_{submission.file_name}"
                        submission_name = f"{question_dir}/{submission_name}"
                        with open(submission_name, "wb") as f:
                            f.write(submission.file)

                    saved_students.add(student.id)

            course_report = f"{course.name}_{course.id}_report"
            course_report = ReportExporter._sanitize(course_report) + ".xlsx"
            report.save(f"{course_dir}/{course_report}")

    def _compress_and_remove_report_dir(self):
        shutil.make_archive(self._output_path.rstrip(
            ".zip"), 'zip', self._report_dir)
        shutil.rmtree(self._report_dir)


class Command(BaseCommand):
    help = 'Export a report of the activities of all courses in a zip file.'

    def add_arguments(self, parser):
        parser.add_argument("output_file",
                            type=str,
                            help='Path to the output file.')

    def handle(self, *args, **options):
        exporter = ReportExporter(options['output_file'])
        exporter.export()
